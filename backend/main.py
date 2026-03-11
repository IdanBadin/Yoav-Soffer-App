"""
FastAPI backend — thin wrapper around existing utils.
Endpoints: POST /process, GET/POST/PUT/DELETE /prices, GET /health, POST /refresh-prices
"""

import asyncio
import base64
import io
import json
import logging
import os
import re
import tempfile
from contextlib import asynccontextmanager
from datetime import datetime
from typing import Optional

from pdf2image import convert_from_bytes as _pdf2image

import anthropic as _anthropic
import gspread
from dotenv import load_dotenv
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from google.oauth2.service_account import Credentials
from pydantic import BaseModel

from utils.pdf_parser import parse_pdf
from utils.sheets_client import build_price_index, load_price_sheet, match_prices
from utils.excel_generator import generate_quote, generate_parts_list

load_dotenv()
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

_price_index: Optional[dict] = None
_price_records_raw: Optional[list] = None  # Original records for semantic matching

READ_SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets.readonly",
    "https://www.googleapis.com/auth/drive.readonly",
]
WRITE_SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]


def _get_credentials_path() -> Optional[str]:
    b64 = os.getenv("GOOGLE_CREDENTIALS_B64", "")
    if b64:
        tmp = tempfile.NamedTemporaryFile(suffix=".json", delete=False, mode="w")
        tmp.write(base64.b64decode(b64).decode("utf-8"))
        tmp.close()
        return tmp.name
    local_path = "config/google_credentials.json"
    if os.path.exists(local_path):
        return local_path
    return None


def _get_worksheet(write: bool = False):
    """Get gspread worksheet with read or write access."""
    creds_path = _get_credentials_path()
    sheet_id = os.getenv("GOOGLE_SHEET_ID", "")
    sheet_name = os.getenv("GOOGLE_SHEET_NAME", "מחירון")
    if not creds_path or not sheet_id:
        raise HTTPException(status_code=500, detail="Google Sheets לא מוגדר")
    scopes = WRITE_SCOPES if write else READ_SCOPES
    creds = Credentials.from_service_account_file(creds_path, scopes=scopes)
    gc = gspread.authorize(creds)
    ss = gc.open_by_key(sheet_id)
    try:
        return ss.worksheet(sheet_name)
    except gspread.WorksheetNotFound:
        return ss.get_worksheet(0)


def _load_price_index() -> Optional[dict]:
    global _price_records_raw
    creds_path = _get_credentials_path()
    sheet_id = os.getenv("GOOGLE_SHEET_ID", "")
    sheet_name = os.getenv("GOOGLE_SHEET_NAME", "מחירון")
    if not creds_path or not sheet_id:
        logger.warning("Google Sheets not configured — prices will be 0")
        return None
    try:
        records = load_price_sheet(creds_path, sheet_id, sheet_name)
        _price_records_raw = records
        index = build_price_index(records)
        logger.info(f"Price index loaded: {len(records)} records")
        return index
    except Exception as e:
        logger.error(f"Failed to load price sheet: {e}")
        return None


def _semantic_match_unmatched(
    unmatched: list,   # [(original_index, component), ...]
    raw_records: list,
    api_key: str,
) -> dict:            # {original_index: {price, unit, match_type, price_found}}
    """Send unmatched components to Claude for cross-language semantic matching."""
    import json, re

    # Build compact price list string
    price_lines = []
    for i, r in enumerate(raw_records):
        cat = r.get("catalog_number", "") or ""
        name = r.get("item_name", "") or ""
        category = r.get("category", "") or ""
        price = r.get("unit_price", 0) or 0
        price_lines.append(f'[{i}] cat:"{cat}" "{name}" [{category}] ₪{price}')
    price_list_str = "\n".join(price_lines)

    # Build compact components string
    comp_lines = []
    for local_i, (orig_i, c) in enumerate(unmatched):
        cat = c.get("catalog_number", "") or ""
        mfg = c.get("manufacturer", "") or ""
        desc = c.get("description", "") or ""
        comp_lines.append(f'[{local_i}] cat:"{cat}" mfg:"{mfg}" "{desc}"')
    comp_str = "\n".join(comp_lines)

    prompt = f"""You are an expert in electrical panels and components used in Israel.

Match each schematic component to the best price list item.
Component names may be in English; price list names are in Hebrew.
Common mappings: Circuit Breaker=מאמת, Residual Current=פחת/מאזם, Contactor=מגע, Motor Protection=הגנת מנוע, Selector=בורר, Timer=טיימר, Relay=ממסר, Capacitor=קבל, Current Transformer=משנז"ר, Surge Protection=כולא ברק

PRICE LIST (index | catalog | name | category | price):
{price_list_str}

COMPONENTS TO MATCH (index | catalog | manufacturer | description):
{comp_str}

Rules:
1. Match by component TYPE + SPECIFICATIONS (poles, amperage, mA rating)
2. Only return matches with HIGH confidence (>85%)
3. For no confident match → use -1

Return ONLY valid JSON, no text:
{{"m": [[0, 15], [1, 22], [2, -1]]}}
where each pair is [component_index, price_list_index] (-1 = no match)"""

    client = _anthropic.Anthropic(api_key=api_key)
    response = client.messages.create(
        model="claude-sonnet-4-5",
        max_tokens=4096,
        messages=[{"role": "user", "content": prompt}],
    )
    raw = response.content[0].text.strip()

    # Use raw_decode to find the first JSON object containing "m" key.
    # raw_decode parses exactly ONE value and stops, so trailing text / extra
    # objects never bleed into the parse (avoids "Extra data" JSONDecodeError).
    decoder = json.JSONDecoder()
    data = None
    pos = 0
    while pos < len(raw):
        try:
            brace_pos = raw.index('{', pos)
            candidate, _ = decoder.raw_decode(raw[brace_pos:])
            if isinstance(candidate, dict) and "m" in candidate:
                data = candidate
                break
            pos = brace_pos + 1
        except (ValueError, json.JSONDecodeError):
            break
    if data is None:
        raise ValueError(f"No JSON with 'm' key found in Claude response: {raw[:200]}")

    result = {}
    for pair in data.get("m", []):
        local_i, price_list_i = pair[0], pair[1]
        if price_list_i == -1:
            continue
        if local_i < 0 or local_i >= len(unmatched):
            continue
        if price_list_i < 0 or price_list_i >= len(raw_records):
            continue
        orig_i = unmatched[local_i][0]
        rec = raw_records[price_list_i]
        try:
            price_val = float(str(rec.get("unit_price", 0) or 0).replace(",", "."))
        except (ValueError, TypeError):
            price_val = 0.0
        result[orig_i] = {
            "price": price_val,
            "unit": str(rec.get("unit", "יח'") or "יח'"),
            "match_type": "semantic",
            "price_found": True,
        }
    return result


# ── BoQ (Bill of Quantities) helpers ──────────────────────────────────────────

def _detect_boq_structure(ws) -> dict:
    """
    Detect header row and column positions in a BoQ worksheet.
    Returns:
        header_row  (int, 1-indexed)
        col_code    (int|None, 0-indexed)
        col_desc    (int, 0-indexed)
        col_unit    (int|None, 0-indexed)
        col_qty     (int, 0-indexed)
        col_price   (int, 0-indexed)
        col_total   (int|None, 0-indexed)
    """
    DESC_KW  = {'תיאור', 'פירוט', 'description', 'שם'}
    QTY_KW   = {'כמות', 'qty', 'quantity'}
    PRICE_KW = {'מחיר', 'price'}
    UNIT_KW  = {"יח'", 'יחידה', 'unit', "יח' מידה", "יח"}
    TOTAL_KW = {'סהכ', 'סה"כ', 'סה כ', 'total'}
    CODE_KW  = {'סעיף', 'מספר', 'קוד', 'code', 'number'}

    header_row = None
    header_vals = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), start=1):
        vals_lower = [str(v).strip().lower() if v is not None else '' for v in row]
        hits = 0
        for v in vals_lower:
            if any(kw.lower() in v for kw in QTY_KW | PRICE_KW | DESC_KW):
                hits += 1
        if hits >= 2:
            header_row = row_idx
            header_vals = [str(v).strip() if v is not None else '' for v in row]
            break

    if header_row is None:
        header_row = 1
        header_vals = [str(v).strip() if v is not None else ''
                       for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

    col_desc = col_qty = col_price = None
    col_code = col_unit = col_total = None

    for i, v in enumerate(header_vals):
        vl = v.lower()
        if col_code  is None and any(k.lower() in vl for k in CODE_KW):  col_code  = i
        if col_desc  is None and any(k.lower() in vl for k in DESC_KW):  col_desc  = i
        if col_unit  is None and any(k.lower() in vl for k in UNIT_KW):  col_unit  = i
        if col_qty   is None and any(k.lower() in vl for k in QTY_KW):   col_qty   = i
        if col_price is None and any(k.lower() in vl for k in PRICE_KW): col_price = i
        if col_total is None and any(k.lower() in vl for k in TOTAL_KW): col_total = i

    # Positional fallback for formats where keywords don't appear in header
    # Observed: [code(0), desc(1), unit(2), qty(3), price(4), total(5)]
    if col_desc  is None: col_desc  = 1
    if col_qty   is None: col_qty   = 3
    if col_price is None: col_price = 4
    if col_unit  is None: col_unit  = 2
    if col_total is None and ws.max_column >= 6: col_total = 5

    return {
        'header_row': header_row,
        'col_code':   col_code,
        'col_desc':   col_desc,
        'col_unit':   col_unit,
        'col_qty':    col_qty,
        'col_price':  col_price,
        'col_total':  col_total,
    }


def _extract_boq_items(ws, structure: dict) -> list:
    """
    Extract priceable component rows from a BoQ worksheet.
    Returns list of dicts with row metadata for writing back and matching.
    """
    SKIP_UNITS = {'הערה', 'note', 'notes', ''}
    items = []
    data_start = structure['header_row'] + 1

    c_desc  = structure['col_desc']
    c_qty   = structure['col_qty']
    c_unit  = structure['col_unit']
    c_code  = structure['col_code']
    c_price = structure['col_price']
    c_total = structure['col_total']

    max_needed = max(v for v in [c_desc, c_qty, c_unit or 0, c_code or 0, c_price] if v is not None)

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=data_start, values_only=True), start=data_start
    ):
        if len(row) <= max_needed:
            continue

        desc    = row[c_desc] if c_desc is not None else None
        qty_raw = row[c_qty]  if c_qty  is not None else None
        unit_raw = row[c_unit] if c_unit is not None else ''
        code_raw = row[c_code] if c_code is not None else ''

        if desc is None or str(desc).strip() == '':
            continue

        desc_str = str(desc).strip()
        unit_str = str(unit_raw).strip() if unit_raw is not None else ''

        if unit_str in SKIP_UNITS or len(desc_str) < 3:
            continue

        try:
            qty = float(str(qty_raw).replace(',', '.').strip()) if qty_raw is not None else 0.0
        except (ValueError, TypeError):
            qty = 0.0

        if qty <= 0:
            continue

        code_str = str(code_raw).strip() if code_raw is not None else ''

        items.append({
            'row_idx':        row_idx,
            'description':    desc_str,
            'qty':            qty,
            'unit':           unit_str or "יח'",
            'code':           code_str,
            'catalog_number': code_str,
            'manufacturer':   '',
            'desc_col':       c_desc,        # 0-indexed, for highlight
            'price_col':      c_price,       # 0-indexed, for writing price
            'total_col':      c_total,       # 0-indexed or None
        })

    return items


def _fill_boq_excel(file_bytes: bytes, items: list, matched: list) -> bytes:
    """
    Write matched prices back into the Excel file in-place.
    Unmatched items get yellow highlight on their description cell.
    Returns modified Excel bytes.
    """
    import io as _io
    from openpyxl import load_workbook as _load_wb
    from openpyxl.styles import PatternFill

    YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    wb = _load_wb(_io.BytesIO(file_bytes))
    ws = wb.active

    for item, match in zip(items, matched):
        row_idx   = item['row_idx']
        price_col = item['price_col'] + 1          # openpyxl is 1-indexed
        desc_col  = item['desc_col'] + 1
        total_col = (item['total_col'] + 1) if item['total_col'] is not None else None

        if match.get('price_found') and match.get('price', 0) > 0:
            ws.cell(row=row_idx, column=price_col).value = match['price']
            if total_col:
                ws.cell(row=row_idx, column=total_col).value = item['qty'] * match['price']
        else:
            ws.cell(row=row_idx, column=desc_col).fill = YELLOW

    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _process_boq_flow(
    file_bytes: bytes,
    price_index: Optional[dict],
    price_records_raw: Optional[list],
    api_key: str,
) -> dict:
    """
    Full BoQ processing pipeline.
    Returns dict with same shape as /process PDF response.
    """
    import io as _io
    from openpyxl import load_workbook as _load_wb

    wb = _load_wb(_io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    structure = _detect_boq_structure(ws)
    items = _extract_boq_items(ws, structure)

    if not items:
        raise ValueError("לא נמצאו פריטים בכתב הכמויות. ודא שהקובץ מכיל עמודות כמות ומחיר.")

    logger.info(f"BoQ: extracted {len(items)} priceable items")

    components_for_match = [
        {
            "description":    item["description"],
            "catalog_number": item["catalog_number"],
            "manufacturer":   item["manufacturer"],
            "qty":            item["qty"],
        }
        for item in items
    ]

    # Steps 1-3: exact + normalized + fuzzy
    if price_index:
        matched = match_prices(components_for_match, price_index)
    else:
        matched = [
            {**c, "price": 0.0, "unit": "יח'", "match_type": "none", "price_found": False}
            for c in components_for_match
        ]

    # Step 4: semantic matching for unmatched
    unmatched = [(i, c) for i, c in enumerate(matched) if not c.get("price_found")]
    if unmatched and price_records_raw and api_key:
        try:
            semantic = _semantic_match_unmatched(unmatched, price_records_raw, api_key)
            for idx, match_data in semantic.items():
                matched[idx] = {**matched[idx], **match_data}
            logger.info(f"BoQ semantic matched {len(semantic)} of {len(unmatched)} unmatched")
        except Exception as e:
            logger.warning(f"BoQ semantic matching skipped: {e}")

    # Merge qty/unit back for display
    for i, item in enumerate(items):
        matched[i] = {
            **matched[i],
            "qty":  item["qty"],
            "unit": matched[i].get("unit") or item["unit"],
        }

    filled_bytes = _fill_boq_excel(file_bytes, items, matched)

    components_display = [
        {
            "description":  item["description"],
            "catalog":      item["catalog_number"],   # frontend Component interface uses 'catalog'
            "user1":        "",
            "manufacturer": matched[i].get("manufacturer", ""),
            "qty":          item["qty"],
            "unit":         matched[i].get("unit", item["unit"]),
            "price":        matched[i].get("price", 0.0),
            "match_type":   matched[i].get("match_type", "none"),
            "price_found":  matched[i].get("price_found", False),
        }
        for i, item in enumerate(items)
    ]

    matched_count = sum(1 for c in components_display if c["price_found"])
    logger.info(f"BoQ: {matched_count}/{len(items)} items priced")

    return {
        "components":  components_display,
        "page_count":  0,
        "excel_quote": base64.b64encode(filled_bytes).decode("utf-8"),
        "excel_parts": "",
        "boq_mode":    True,
    }


def _vision_extract_page(client, img_bytes: bytes, page_num: int) -> list:
    """Send a single PDF page image to Claude Vision and return extracted components."""
    b64 = base64.standard_b64encode(img_bytes).decode()
    prompt = (
        "You are analyzing an AutoCAD electrical panel (לוח חשמל) drawing.\n\n"
        "Extract ONLY purchasable electrical components — physical items that appear in a bill of materials.\n\n"
        "INCLUDE: circuit breakers (מפסק), contactors (מגען), motor protection relays (מגן מנוע), "
        "residual current devices (מפסק פחת/מגן מנוע), terminals (מסוף), busbars (שפה/בסבר), "
        "meters (מד), pilot lights (נורה), selector switches (בורר), timers (טיימר), "
        "transformers (שנאי/טרנספורמטור), fuses (פיוז), cables (כבל), surge protectors (מוגן ברק), "
        "enclosures (ארון חשמל/לוח), din rails, cable ducts (מרזב).\n\n"
        "EXCLUDE: specifications text, dimensions, weights, standards (IEC/EN), notes, dates, "
        "temperature/IP ratings, company names, drawing titles, section headers, calculations.\n\n"
        "For each component return a JSON object:\n"
        '  "description": string — component name + key specs (type, poles, rating)\n'
        '  "qty": number — integer quantity (default 1)\n'
        '  "unit": string — use "יח\'" for pieces, "מ\'" for meters\n'
        '  "catalog": string — model/catalog number if visible, else ""\n\n'
        "Return ONLY a JSON array, no markdown, no explanation. Example:\n"
        '[{"description":"מפסק אוטומטי 3P 63A","qty":2,"unit":"יח\'","catalog":""},\n'
        ' {"description":"מגן מנוע 11-16A","qty":1,"unit":"יח\'","catalog":"GV2ME16"}]\n\n'
        "If no purchasable components are visible on this page, return []."
    )
    try:
        msg = client.messages.create(
            model="claude-sonnet-4-5",
            max_tokens=4096,
            messages=[{
                "role": "user",
                "content": [
                    {"type": "image", "source": {"type": "base64", "media_type": "image/png", "data": b64}},
                    {"type": "text", "text": prompt},
                ],
            }],
        )
        text = msg.content[0].text.strip()
        # Strip markdown code fences if present
        text = re.sub(r"^```[a-z]*\n?", "", text)
        text = re.sub(r"\n?```$", "", text)
        # Extract just the JSON array
        json_match = re.search(r'\[.*\]', text, re.DOTALL)
        if not json_match:
            logger.warning(f"Vision page {page_num}: no JSON array found in response")
            return []
        arr_text = json_match.group()
        try:
            return json.loads(arr_text)
        except json.JSONDecodeError:
            # Fix invalid escape sequences Claude sometimes emits (e.g. \4, \מ)
            arr_fixed = re.sub(r'\\(?!["\\/bfnrtu])', r'\\\\', arr_text)
            return json.loads(arr_fixed)
    except Exception as e:
        logger.warning(f"Vision page {page_num} extraction failed: {e}")
        return []


async def _process_pdf_vision(
    file_bytes: bytes,
    price_index: Optional[dict],
    price_records_raw: Optional[list],
    api_key: str,
    project_name: str,
    manager_name: str,
    date: str,
) -> dict:
    """
    Vision API PDF pipeline: pdf2image → Claude Vision per page (parallel) → dedup → price match.
    Returns same shape as the old PDF /process response.
    """
    client = _anthropic.Anthropic(api_key=api_key)

    # Convert PDF pages to PNG images at 150 DPI
    try:
        images = _pdf2image(file_bytes, dpi=150, fmt="png")
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"שגיאה בפענוח PDF: {e}")

    page_count = len(images)
    logger.info(f"Vision PDF: {page_count} page(s) — processing in parallel...")

    # Claude Vision max image dimension is 8000px — resize if needed
    MAX_DIM = 7500

    def _page_bytes(img) -> bytes:
        w, h = img.size
        if max(w, h) > MAX_DIM:
            scale = MAX_DIM / max(w, h)
            img = img.resize((int(w * scale), int(h * scale)))
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        return buf.getvalue()

    # Process all pages concurrently via asyncio.to_thread
    async def _extract_page_async(img, page_num: int) -> list:
        img_bytes = _page_bytes(img)
        items = await asyncio.to_thread(_vision_extract_page, client, img_bytes, page_num)
        logger.info(f"  Page {page_num}: {len(items)} components extracted")
        return items

    page_results = await asyncio.gather(*[
        _extract_page_async(img, i + 1) for i, img in enumerate(images)
    ])
    all_items: list = [item for page in page_results for item in page]

    # Deduplicate: merge items with identical description by summing qty
    merged: dict = {}
    for item in all_items:
        key = (item.get("description") or "").strip().lower()
        if not key:
            continue
        if key in merged:
            merged[key]["qty"] = merged[key].get("qty", 1) + item.get("qty", 1)
        else:
            merged[key] = {
                "description": item.get("description", ""),
                "qty":         item.get("qty", 1),
                "unit":        item.get("unit", "יח'"),
                "catalog":     item.get("catalog", ""),
            }

    components_raw = list(merged.values())
    logger.info(f"Vision PDF: {len(components_raw)} unique components after dedup")

    if not components_raw:
        raise HTTPException(
            status_code=422,
            detail="לא זוהו רכיבים בשרטוט. ודא שהקובץ הנכון הועלה ושהוא מכיל רשימת ציוד."
        )

    # Build format expected by match_prices()
    comps_for_match = [
        {
            "description":    c["description"],
            "catalog_number": c["catalog"],
            "manufacturer":   "",
            "qty":            c["qty"],
            "unit":           c["unit"],
            "user1":          "",
        }
        for c in components_raw
    ]

    # Steps 1-3: exact + normalized + fuzzy
    if price_index:
        matched = match_prices(comps_for_match, price_index)
    else:
        matched = [
            {**c, "price": 0.0, "match_type": "none", "price_found": False}
            for c in comps_for_match
        ]

    # Step 4: semantic matching for unmatched (batch of 80 to keep prompt manageable)
    unmatched = [(i, c) for i, c in enumerate(matched) if not c.get("price_found")]
    SEMANTIC_BATCH = 80
    if unmatched and price_records_raw and api_key:
        total_sem = 0
        for batch_start in range(0, len(unmatched), SEMANTIC_BATCH):
            batch = unmatched[batch_start:batch_start + SEMANTIC_BATCH]
            try:
                semantic = _semantic_match_unmatched(batch, price_records_raw, api_key)
                # semantic keys are orig_i (index into matched) — already converted inside the function
                for orig_idx, upd in semantic.items():
                    matched[orig_idx] = {**matched[orig_idx], **upd}
                total_sem += len(semantic)
            except Exception as e:
                logger.warning(f"Vision semantic matching batch skipped: {e}")
        logger.info(f"Vision semantic matched {total_sem} of {len(unmatched)} unmatched")

    # Restore qty/unit from original extraction (match_prices may overwrite unit)
    for i, raw in enumerate(components_raw):
        matched[i]["qty"]  = raw["qty"]
        matched[i]["unit"] = matched[i].get("unit") or raw["unit"]

    excel_quote_bytes = generate_quote(matched, project_name, manager_name, date)
    excel_parts_bytes = generate_parts_list(matched, project_name, manager_name, date)

    matched_count = sum(1 for c in matched if c.get("price_found"))
    logger.info(f"Vision PDF: {matched_count}/{len(matched)} components priced")

    return {
        "components": matched,
        "page_count":  page_count,
        "excel_quote": base64.b64encode(excel_quote_bytes).decode("utf-8"),
        "excel_parts": base64.b64encode(excel_parts_bytes).decode("utf-8"),
    }


@asynccontextmanager
async def lifespan(app: FastAPI):
    global _price_index
    _price_index = _load_price_index()
    yield


app = FastAPI(title="י. סופר — Quote Automation API", lifespan=lifespan)

allowed_origins = os.getenv("ALLOWED_ORIGINS", "*").split(",")
app.add_middleware(
    CORSMiddleware,
    allow_origins=allowed_origins,
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    allow_headers=["*"],
)


class PriceRecord(BaseModel):
    catalog_number: str = ""
    item_name: str = ""
    unit_price: float = 0.0
    unit: str = "יח'"
    manufacturer: str = ""
    category: str = ""
    cost: str = ""
    notes: str = ""


@app.get("/health")
async def health():
    count = len(_price_index.get("records", [])) if _price_index else 0
    return {
        "status": "ok",
        "prices_loaded": _price_index is not None,
        "price_count": count,
        "timestamp": datetime.now().isoformat(),
    }


@app.post("/refresh-prices")
async def refresh_prices():
    global _price_index
    _price_index = _load_price_index()
    count = len(_price_index.get("records", [])) if _price_index else 0
    return {"status": "ok", "prices_loaded": _price_index is not None, "count": count}


# ── Price CRUD ────────────────────────────────────────────────────────────────

@app.get("/prices")
async def get_prices():
    """Return all rows from the Google Sheet price list."""
    try:
        ws = _get_worksheet(write=False)
        records = ws.get_all_records()
        result = []
        for i, r in enumerate(records, start=2):
            try:
                price_val = float(str(r.get("unit_price", 0) or 0).replace(",", "."))
            except (ValueError, TypeError):
                price_val = 0.0
            result.append({
                "row": i,
                "catalog_number": str(r.get("catalog_number", "") or ""),
                "item_name": str(r.get("item_name", "") or ""),
                "unit_price": price_val,
                "unit": str(r.get("unit", "יח'") or "יח'"),
                "manufacturer": str(r.get("manufacturer", "") or ""),
                "category": str(r.get("category", "") or ""),
                "cost": str(r.get("cost", "") or ""),
                "notes": str(r.get("notes", "") or ""),
            })
        return {"records": result, "count": len(result)}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching prices: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="שגיאה בטעינת המחירון")


@app.put("/prices/{row}")
async def update_price(row: int, data: PriceRecord):
    """Update an existing price row by spreadsheet row number."""
    try:
        ws = _get_worksheet(write=True)
        headers = ws.row_values(1)
        col_map = {h.lower().strip(): idx + 1 for idx, h in enumerate(headers)}
        values = {
            "catalog_number": data.catalog_number,
            "item_name": data.item_name,
            "unit_price": data.unit_price,
            "unit": data.unit,
            "manufacturer": data.manufacturer,
            "category": data.category,
            "cost": data.cost,
            "notes": data.notes,
        }
        cell_updates = []
        for field, col in col_map.items():
            if field in values:
                cell_updates.append({
                    "range": gspread.utils.rowcol_to_a1(row, col),
                    "values": [[values[field]]],
                })
        if cell_updates:
            ws.batch_update(cell_updates)
        return {"status": "ok"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating price row {row}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="שגיאה בעדכון המחיר")


@app.post("/prices")
async def add_price(data: PriceRecord):
    """Append a new price row to the sheet."""
    try:
        ws = _get_worksheet(write=True)
        ws.append_row([
            data.catalog_number,
            data.item_name,
            data.unit_price,
            data.unit,
            data.manufacturer,
            data.category,
            data.cost,
            data.notes,
        ])
        return {"status": "ok"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error adding price: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="שגיאה בהוספת פריט")


@app.delete("/prices/{row}")
async def delete_price(row: int):
    """Delete a price row by spreadsheet row number."""
    try:
        ws = _get_worksheet(write=True)
        ws.delete_rows(row)
        return {"status": "ok"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting price row {row}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="שגיאה במחיקת הפריט")


# ── Main processing endpoint ──────────────────────────────────────────────────

@app.post("/process")
async def process(
    file: UploadFile = File(...),
    project_name: str = Form(...),
    manager_name: str = Form(""),
    date: str = Form(...),
):
    api_key = os.getenv("ANTHROPIC_API_KEY", "")
    if not api_key:
        raise HTTPException(
            status_code=500,
            detail="ANTHROPIC_API_KEY לא מוגדר בסביבת הייצור. יש להגדיר אותו ב-Railway Variables."
        )

    fname_lower = (file.filename or "").lower()
    is_pdf  = fname_lower.endswith(".pdf")
    is_xlsx = fname_lower.endswith(".xlsx") or fname_lower.endswith(".xls")

    if not is_pdf and not is_xlsx:
        raise HTTPException(
            status_code=422,
            detail="יש להעלות קובץ PDF (שרטוט) או Excel (כתב כמויות). סוג הקובץ שהועלה אינו נתמך."
        )

    MAX_SIZE = 50 * 1024 * 1024  # 50 MB
    tmp_path = None
    try:
        content = await file.read()

        if len(content) == 0:
            raise HTTPException(status_code=422, detail="הקובץ שהועלה ריק. ודא שהקובץ תקין ונסה שנית.")

        if len(content) > MAX_SIZE:
            size_mb = len(content) / (1024 * 1024)
            raise HTTPException(
                status_code=422,
                detail=f"הקובץ גדול מדי ({size_mb:.1f} MB). הגודל המקסימלי הוא 50 MB."
            )

        # ── Excel BoQ flow ────────────────────────────────────────────────────
        if is_xlsx:
            logger.info(f"Processing BoQ Excel: {file.filename} ({len(content) / 1024:.0f} KB)")
            return _process_boq_flow(content, _price_index, _price_records_raw, api_key)

        # ── PDF → Vision API flow ─────────────────────────────────────────────
        logger.info(f"Processing PDF (Vision): {file.filename} ({len(content) / 1024:.0f} KB)")
        return await _process_pdf_vision(
            content, _price_index, _price_records_raw, api_key,
            project_name, manager_name, date,
        )

    except HTTPException:
        raise
    except _anthropic.AuthenticationError:
        logger.error("Anthropic authentication failed")
        raise HTTPException(
            status_code=401,
            detail="מפתח ה-API של Anthropic אינו תקין. יש לעדכן את ANTHROPIC_API_KEY ב-Railway Variables."
        )
    except _anthropic.RateLimitError:
        logger.error("Anthropic rate limit hit")
        raise HTTPException(
            status_code=429,
            detail="חרגת ממגבלת הקריאות ל-Claude AI. המתן מספר שניות ונסה שנית."
        )
    except _anthropic.APIConnectionError as e:
        logger.error(f"Anthropic connection error: {e}")
        raise HTTPException(
            status_code=503,
            detail="לא ניתן להתחבר ל-Claude AI. ייתכן ויש בעיית רשת זמנית — נסה שנית."
        )
    except _anthropic.BadRequestError as e:
        logger.error(f"Anthropic bad request: {e}")
        raise HTTPException(
            status_code=422,
            detail="קובץ ה-PDF מכיל יותר מדי נתונים לעיבוד בבת אחת. נסה לפצל את הקובץ לחלקים."
        )
    except MemoryError:
        logger.error("MemoryError during processing")
        raise HTTPException(
            status_code=500,
            detail="קובץ ה-PDF גדול מדי לעיבוד. נסה קובץ קטן יותר."
        )
    except Exception as e:
        logger.error(f"Processing error: {e}", exc_info=True)
        err = str(e)
        if "credit balance" in err.lower() or "too low" in err.lower():
            raise HTTPException(
                status_code=402,
                detail="יתרת הקרדיטים ב-Anthropic אזלה. יש להוסיף קרדיטים בכתובת console.anthropic.com"
            )
        if "could not be decoded" in err.lower() or "unable to" in err.lower() or "pdf" in err.lower():
            raise HTTPException(
                status_code=422,
                detail="קובץ ה-PDF פגום או מוצפן ולא ניתן לקרוא אותו. נסה לייצא את הקובץ מחדש."
            )
        raise HTTPException(status_code=500, detail=f"שגיאה בעיבוד השרטוט: {err[:200]}")
    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.unlink(tmp_path)
