"""
Excel Generator — produces pixel-perfect Excel files matching Y. Sofer's format.

FILE 1: הצעת מחיר/חשבון (Quote/Invoice) — 5-column format
FILE 2: כתב חלקים (Parts List) — 8-column BOM format

All measurements confirmed from sample file analysis:
  - Column widths, row heights, fonts (Calibri Light), colors, borders, RTL
"""

import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    GradientFill,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ─── Color constants (confirmed from sample analysis) ────────────────────────
BLUE_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")   # theme:3 deep navy-blue
YELLOW_WARN_FILL = PatternFill("solid", fgColor="FFF2CC")   # amber for unmatched prices
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
NO_FILL = PatternFill(fill_type=None)

# ─── Font constants ───────────────────────────────────────────────────────────
def font(size=12, bold=False, color="000000", name="Calibri Light"):
    return Font(name=name, size=size, bold=bold, color=color)

HEADER_FONT = font(size=18, bold=True, color="FFFFFF")
COL_HEADER_FONT = font(size=12, bold=True, color="FFFFFF")
DATA_FONT = font(size=12)
BOLD_FONT = font(size=12, bold=True)
META_FONT = font(size=11, name="Calibri")
META_LABEL_FONT = font(size=11, bold=True, name="Calibri")

# ─── Border constants ─────────────────────────────────────────────────────────
THIN = Side(style="thin")
MEDIUM = Side(style="medium")
THICK = Side(style="thick")

def border(left=None, right=None, top=None, bottom=None):
    return Border(left=left or Side(), right=right or Side(),
                  top=top or Side(), bottom=bottom or Side())

BORDER_ALL_THIN = border(THIN, THIN, THIN, THIN)
BORDER_LEFT_MED_RIGHT_THIN = border(MEDIUM, THIN, THIN, THIN)
BORDER_LEFT_MED_RIGHT_MED = border(MEDIUM, MEDIUM, THIN, THIN)
BORDER_TOTAL_ROW = border(MEDIUM, MEDIUM, THIN, MEDIUM)
BORDER_HEADER_LEFT = border(MEDIUM, THIN, MEDIUM, THIN)
BORDER_HEADER_RIGHT = border(THIN, MEDIUM, MEDIUM, THIN)
BORDER_HEADER_MID = border(THIN, THIN, MEDIUM, THIN)

# ─── Alignment constants ──────────────────────────────────────────────────────
def align(h="right", v="center", wrap=False, rtl=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap,
                     readingOrder=2 if rtl else 1)

CENTER = align(h="center")
CENTER_WRAP = align(h="center", wrap=True)
RIGHT = align(h="right")
RIGHT_WRAP = align(h="right", wrap=True)


# ─────────────────────────────────────────────────────────────────────────────
# FILE 1 — הצעת מחיר / חשבון
# ─────────────────────────────────────────────────────────────────────────────

def generate_quote(
    components: list[dict],
    project_name: str,
    manager_name: str,
    quote_date: str,
) -> bytes:
    """
    Generate File 1 — הצעת מחיר / חשבון.
    Format confirmed from: דוגמא לפורמט מלא של הצעת מחיר או חשבון ללקוח.xlsx

    Columns: A=סעיף | B=הערה | C=כמות | D=מחיר | E=סה"כ
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "הצעת מחיר"
    ws.sheet_view.rightToLeft = True

    # ── Column widths (confirmed from sample) ────────────────────────────────
    ws.column_dimensions["A"].width = 35.375
    ws.column_dimensions["B"].width = 11.25
    ws.column_dimensions["C"].width = 12.5
    ws.column_dimensions["D"].width = 12.125
    ws.column_dimensions["E"].width = 12.875

    # ── Rows 1-3: Header block ───────────────────────────────────────────────
    for r in range(1, 4):
        ws.row_dimensions[r].height = 15.75

    # Company name block (A1:B3 merged)
    ws.merge_cells("A1:B3")
    ws["A1"] = "י. סופר מערכות חשמל"
    ws["A1"].font = font(size=13, bold=True, name="Calibri")
    ws["A1"].alignment = align(h="center")

    # Meta labels (col C) and values (col D-E merged)
    meta = [
        ("C1", "תאריך:", "D1", "E1", quote_date),
        ("C2", "פרויקט:", "D2", "E2", project_name),
        ("C3", "מנהל פרויקט:", "D3", "E3", manager_name),
    ]
    for label_cell, label_text, val_start, val_end, val_text in meta:
        ws[label_cell] = label_text
        ws[label_cell].font = META_LABEL_FONT
        ws[label_cell].alignment = align(h="right")
        ws.merge_cells(f"{val_start}:{val_end}")
        ws[val_start] = val_text
        ws[val_start].font = META_FONT
        ws[val_start].alignment = align(h="right")

    # ── Row 4: Title bar ──────────────────────────────────────────────────────
    ws.row_dimensions[4].height = 23.25
    ws.merge_cells("A4:E4")
    ws["A4"] = f"הצעת מחיר - {project_name}"
    ws["A4"].font = HEADER_FONT
    ws["A4"].fill = BLUE_HEADER_FILL
    ws["A4"].alignment = CENTER
    ws["A4"].border = border(MEDIUM, MEDIUM, MEDIUM, THIN)

    # ── Row 5: Column headers ─────────────────────────────────────────────────
    ws.row_dimensions[5].height = 22.5
    headers = ["סעיף", "הערה", "כמות", "מחיר", 'סה"כ']
    for col_idx, header_text in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx, value=header_text)
        cell.font = COL_HEADER_FONT
        cell.fill = BLUE_HEADER_FILL
        cell.alignment = CENTER
        if col_idx == 1:
            cell.border = border(MEDIUM, THIN, THIN, THIN)
        elif col_idx == 5:
            cell.border = border(THIN, MEDIUM, THIN, THIN)
        else:
            cell.border = BORDER_ALL_THIN

    # ── Data rows ─────────────────────────────────────────────────────────────
    data_start = 6
    for i, comp in enumerate(components):
        row = data_start + i
        ws.row_dimensions[row].height = 15.75

        description = comp.get("description", "")
        unit = comp.get("unit", "יח'")
        qty = comp.get("qty", 0)
        price = comp.get("price", 0)
        price_found = comp.get("price_found", True)

        fill = YELLOW_WARN_FILL if not price_found else NO_FILL

        def _cell(col, value, h_align="center", wrap=False):
            c = ws.cell(row=row, column=col, value=value)
            c.font = DATA_FONT
            c.fill = fill
            c.alignment = align(h=h_align, wrap=wrap)
            return c

        c_a = _cell(1, description, h_align="right", wrap=True)
        c_a.border = border(MEDIUM, THIN, THIN, THIN)

        c_b = _cell(2, unit)
        c_b.border = BORDER_ALL_THIN

        c_c = _cell(3, qty)
        c_c.border = BORDER_ALL_THIN

        c_d = _cell(4, price if price else None)
        c_d.border = BORDER_ALL_THIN

        # Formula: =D*C
        d_col = get_column_letter(4)
        c_col = get_column_letter(3)
        c_e = ws.cell(row=row, column=5)
        c_e.value = f"={d_col}{row}*{c_col}{row}"
        c_e.font = DATA_FONT
        c_e.fill = fill
        c_e.alignment = CENTER
        c_e.border = border(THIN, MEDIUM, THIN, THIN)

    # ── Total row ─────────────────────────────────────────────────────────────
    total_row = data_start + len(components)
    ws.row_dimensions[total_row].height = 26.25

    last_data_row = total_row - 1

    total_label = ws.cell(row=total_row, column=1, value='סה"כ')
    total_label.font = BOLD_FONT
    total_label.fill = WHITE_FILL
    total_label.alignment = CENTER
    total_label.border = BORDER_TOTAL_ROW

    for col in range(2, 5):
        c = ws.cell(row=total_row, column=col)
        c.fill = WHITE_FILL
        c.border = BORDER_TOTAL_ROW

    total_sum = ws.cell(row=total_row, column=5)
    total_sum.value = f"=SUM(E{data_start}:E{last_data_row})"
    total_sum.font = BOLD_FONT
    total_sum.fill = WHITE_FILL
    total_sum.alignment = CENTER
    total_sum.border = border(THIN, MEDIUM, THIN, MEDIUM)

    # ── Output as BytesIO ─────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────────────────────────
# FILE 2 — כתב חלקים (Parts List)
# ─────────────────────────────────────────────────────────────────────────────

def generate_parts_list(
    components: list[dict],
    project_name: str,
    manager_name: str,
    quote_date: str,
) -> bytes:
    """
    Generate File 2 — כתב חלקים (detailed BOM parts list).
    8 columns: מס' | תיאור | מק"ט | יצרן | כמות | יחידה | מחיר יחידה | סה"כ
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "כתב חלקים"
    ws.sheet_view.rightToLeft = True

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = [7, 35, 18, 18, 8, 8, 13, 13]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Rows 1-2: Header block ────────────────────────────────────────────────
    ws.row_dimensions[1].height = 15.75
    ws.row_dimensions[2].height = 15.75

    ws.merge_cells("A1:H1")
    ws["A1"] = "י. סופר מערכות חשמל — כתב חלקים"
    ws["A1"].font = font(size=14, bold=True, name="Calibri")
    ws["A1"].fill = BLUE_HEADER_FILL
    ws["A1"].alignment = CENTER
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")

    meta_cells = [
        ("A2", f"פרויקט: {project_name}"),
        ("D2", f"מנהל: {manager_name}"),
        ("G2", f"תאריך: {quote_date}"),
    ]
    for cell_ref, text in meta_cells:
        ws[cell_ref] = text
        ws[cell_ref].font = META_LABEL_FONT
        ws[cell_ref].alignment = align(h="right")

    # ── Row 3: Column headers ─────────────────────────────────────────────────
    ws.row_dimensions[3].height = 22.5
    col_headers = ["מס'", "תיאור", "מק\"ט", "יצרן", "כמות", "יחידה", "מחיר יחידה", 'סה"כ']
    for col_idx, header_text in enumerate(col_headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header_text)
        cell.font = COL_HEADER_FONT
        cell.fill = BLUE_HEADER_FILL
        cell.alignment = CENTER
        if col_idx == 1:
            cell.border = border(MEDIUM, THIN, THIN, THIN)
        elif col_idx == 8:
            cell.border = border(THIN, MEDIUM, THIN, THIN)
        else:
            cell.border = BORDER_ALL_THIN

    # ── Data rows ─────────────────────────────────────────────────────────────
    data_start = 4
    for i, comp in enumerate(components):
        row = data_start + i
        ws.row_dimensions[row].height = 15.75

        serial = i + 1
        description = comp.get("description", "")
        catalog = comp.get("catalog", "")
        mfg = comp.get("manufacturer", "")
        qty = comp.get("qty", 0)
        unit = comp.get("unit", "יח'")
        price = comp.get("price", 0)
        price_found = comp.get("price_found", True)

        fill = YELLOW_WARN_FILL if not price_found else NO_FILL

        values = [serial, description, catalog, mfg, qty, unit, price if price else None]
        for col_idx, value in enumerate(values, start=1):
            c = ws.cell(row=row, column=col_idx, value=value)
            c.font = DATA_FONT
            c.fill = fill
            c.alignment = RIGHT_WRAP if col_idx == 2 else CENTER
            if col_idx == 1:
                c.border = border(MEDIUM, THIN, THIN, THIN)
            elif col_idx == 7:
                c.border = BORDER_ALL_THIN
            else:
                c.border = BORDER_ALL_THIN

        # Column 8: =G*E (price × qty)
        g_col = get_column_letter(7)
        e_col = get_column_letter(5)
        c_total = ws.cell(row=row, column=8)
        c_total.value = f"={g_col}{row}*{e_col}{row}"
        c_total.font = DATA_FONT
        c_total.fill = fill
        c_total.alignment = CENTER
        c_total.border = border(THIN, MEDIUM, THIN, THIN)

    # ── Total row ─────────────────────────────────────────────────────────────
    total_row = data_start + len(components)
    ws.row_dimensions[total_row].height = 26.25
    last_data_row = total_row - 1

    ws.merge_cells(f"A{total_row}:G{total_row}")
    total_label = ws.cell(row=total_row, column=1, value='סה"כ כולל מע"מ')
    total_label.font = BOLD_FONT
    total_label.fill = WHITE_FILL
    total_label.alignment = CENTER
    total_label.border = BORDER_TOTAL_ROW

    total_sum = ws.cell(row=total_row, column=8)
    total_sum.value = f"=SUM(H{data_start}:H{last_data_row})"
    total_sum.font = BOLD_FONT
    total_sum.fill = WHITE_FILL
    total_sum.alignment = CENTER
    total_sum.border = border(THIN, MEDIUM, THIN, MEDIUM)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
