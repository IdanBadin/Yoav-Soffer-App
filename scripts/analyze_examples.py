"""
analyze_examples.py — One-time script to extract learned knowledge from client example files.

Reads all quote Excel files in client_examples/, extracts component descriptions,
uses Claude to organize vocabulary and build mapping hints, and writes:
  - backend/knowledge/extraction_hints.json  (Vision prompt vocabulary)
  - backend/knowledge/learned_mappings.json  (semantic matching few-shot patterns)

Usage:
  cd /Users/idanbadin/Desktop/Yoav\ Sofer
  ANTHROPIC_API_KEY=<key> python3 scripts/analyze_examples.py

Optional: fetch live price list from production:
  ANTHROPIC_API_KEY=<key> FETCH_PRICES=1 python3 scripts/analyze_examples.py
"""

import os
import sys
import json
import re
import requests
from collections import Counter
from datetime import date

import openpyxl
import anthropic

# ── Paths ──────────────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
EXAMPLES_BASE = os.path.join(PROJECT_ROOT, "client_examples",
    "דוגמאות מוכנות לשרטוטים והצעות מחיר לרפרנס ולמידה")
OUT_DIR = os.path.join(PROJECT_ROOT, "backend", "knowledge")
PRODUCTION_API = "https://yoav-soffer-app-production.up.railway.app"

os.makedirs(OUT_DIR, exist_ok=True)

# ── Step 1: Discover all quote Excel files ─────────────────────────────────

def find_quote_files() -> list[dict]:
    """Walk examples directory and find all quote/invoice Excel files."""
    cases = []
    for case_name in sorted(os.listdir(EXAMPLES_BASE)):
        case_path = os.path.join(EXAMPLES_BASE, case_name)
        if not os.path.isdir(case_path):
            continue
        excels = [
            f for f in os.listdir(case_path)
            if f.endswith(".xlsx") and not f.startswith(".")
        ]
        if not excels:
            continue
        # Prefer חשבון (final invoice) over הצעת מחיר (quote)
        chosen = next((e for e in excels if "חשבון" in e), excels[0])
        cases.append({"case": case_name, "file": os.path.join(case_path, chosen)})
    return cases


# ── Step 2: Extract component rows from quote Excel ────────────────────────

def extract_quote_rows(xlsx_path: str) -> list[dict]:
    """
    Extract all component rows from a quote Excel file.
    Format: A=description | B=unit | C=qty | D=price | E=total
    Skips header rows, metadata rows, and total rows.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    rows = []
    in_data = False

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        a, b, c, d, e = (row[i] if len(row) > i else None for i in range(5))

        # Detect header row (סעיף/description column)
        if isinstance(a, str) and a.strip() in ("סעיף", "description", "פריט"):
            in_data = True
            continue

        if not in_data:
            continue

        # Skip empty rows
        if a is None:
            continue

        desc = str(a).strip()

        # Skip summary/total rows
        skip_patterns = ['סה"כ', "הנחה", "מע\"מ", "Total", "סך הכל", "סה כ"]
        if any(p in desc for p in skip_patterns):
            continue

        # Must have a numeric quantity and price
        try:
            qty = float(c) if c is not None else 0
            price = float(d) if d is not None else 0
        except (TypeError, ValueError):
            continue

        if qty <= 0 or price <= 0:
            continue

        unit = str(b).strip() if b is not None else "יח'"
        rows.append({
            "description": desc,
            "unit": unit,
            "qty": qty,
            "unit_price": price,
        })

    return rows


# ── Step 3: Fetch price list (optional) ───────────────────────────────────

def fetch_price_list() -> list[dict]:
    """Fetch the current price list from the production API."""
    print("  Fetching price list from production API...")
    try:
        resp = requests.get(f"{PRODUCTION_API}/prices", timeout=30)
        resp.raise_for_status()
        data = resp.json()
        print(f"  Got {len(data)} price records.")
        return data
    except Exception as e:
        print(f"  WARNING: Could not fetch price list: {e}")
        return []

def normalize_price_list(raw) -> list[dict]:
    """Handle both list and {records: [...]} response formats."""
    if isinstance(raw, list):
        return raw
    if isinstance(raw, dict):
        return raw.get("records", [])
    return []


# ── Step 4: Use Claude to analyze patterns ─────────────────────────────────

def call_claude(client, prompt: str, max_tokens: int = 4096) -> dict:
    """Call Claude and robustly parse the JSON response. Recovers truncated arrays."""
    msg = client.messages.create(
        model="claude-sonnet-4-5",
        max_tokens=max_tokens,
        messages=[{"role": "user", "content": prompt}],
    )
    text = msg.content[0].text.strip()
    # Strip markdown fences
    text = re.sub(r"^```[a-z]*\n?", "", text, flags=re.MULTILINE)
    text = re.sub(r"\n?```$", "", text)

    # Find outermost { }
    start = text.find('{')
    if start == -1:
        raise ValueError(f"No JSON object in response:\n{text[:300]}")

    end = text.rfind('}')
    candidate = text[start:end+1] if end != -1 else text[start:]

    # Try direct parse first
    try:
        return json.loads(candidate)
    except json.JSONDecodeError:
        pass

    # Truncation recovery: find the last complete array item (ends with }) and close the structure
    # Try truncating at the last complete {...} inside the array
    last_brace = candidate.rfind('}')
    while last_brace > 0:
        truncated = candidate[:last_brace+1]
        # Count open brackets/braces to see if we can close them
        opens = truncated.count('[') - truncated.count(']')
        open_braces = truncated.count('{') - truncated.count('}')
        closing = ']' * opens + '}' * open_braces
        try:
            return json.loads(truncated + closing)
        except json.JSONDecodeError:
            pass
        last_brace = candidate.rfind('}', 0, last_brace)

    pos = 0
    ctx = candidate[max(0, pos - 100):min(len(candidate), pos + 100)]
    raise ValueError(f"JSON parse failed after recovery attempts.\nContext: ...{ctx}...")


def analyze_with_claude(
    all_descriptions: list[str],
    price_list: list[dict],
    api_key: str,
) -> dict:
    """
    Use Claude to analyze client vocabulary in two focused calls:
    1. extraction_hints — component vocabulary for Vision prompt
    2. learned_mappings — description patterns for semantic matching
    """
    client = anthropic.Anthropic(api_key=api_key)

    unique_descs = sorted(set(all_descriptions))
    desc_counts = Counter(all_descriptions)

    # Price list context (up to 100 items for context)
    price_ctx = ""
    if price_list:
        price_ctx = "\nPRICE LIST ITEMS (item_name | category):\n" + "\n".join(
            f"  {r.get('item_name','')} | {r.get('category','')}"
            for r in price_list[:100]
        )

    desc_list = "\n".join(f"[{desc_counts[d]}x] {d}" for d in unique_descs)

    # ── Call 1: extraction_hints ──────────────────────────────────────────
    print("  Call 1/2: building extraction vocabulary...")
    hints_prompt = f"""You are an Israeli electrical panel expert. Analyze these {len(unique_descs)} component descriptions from real quotes by י. סופר מערכות חשמל.
[count] = appearances across 19 projects.

DESCRIPTIONS:
{desc_list}
{price_ctx}

Produce a JSON object for enriching a Vision AI prompt. Keep it compact.

Return ONLY this JSON (no explanation):
{{
  "abbreviations": {{
    "מאמ\\"ת": "main circuit breaker",
    "מא\\"ז": "single/3-pole MCB",
    "ממסר פחת": "RCD relay",
    "כולאי ברק": "surge protector",
    "משנ\\"ז": "current transformer",
    "PKZM": "Moeller motor protection"
  }},
  "component_types": [
    {{"key": "enclosure", "hebrew_terms": ["מבנה לוח", "ארון"], "unit": "מ\\"ר"}},
    {{"key": "main_mcb", "hebrew_terms": ["מאמ\\"ת", "מאמת"], "unit": "יח'"}},
    {{"key": "small_mcb", "hebrew_terms": ["מא\\"ז", "מפסק"], "unit": "יח'"}},
    ... one entry per distinct component category
  ],
  "vision_vocabulary_hint": "2-3 sentence summary of the client vocabulary for Vision prompt injection"
}}"""

    extraction_hints = call_claude(client, hints_prompt, max_tokens=3000)

    # ── Call 2: learned_mappings ──────────────────────────────────────────
    print("  Call 2/2: building semantic mapping patterns...")
    mappings_prompt = f"""You are an Israeli electrical panel expert. For each client description below, provide a search hint for finding the matching item in a Hebrew price list.
[count] = appearances.

DESCRIPTIONS:
{desc_list}
{price_ctx}

Return ONLY this JSON:
{{
  "version": "{date.today().isoformat()}",
  "patterns": [
    {{"pattern": "מא\\"ז עד 1X32A", "type": "single_pole_mcb", "hint": "single-pole circuit breaker up to 32A", "freq": 8}},
    {{"pattern": "מבנה לוח עד 250A", "type": "enclosure", "hint": "panel enclosure/housing rated up to 250A", "freq": 5}},
    ... one entry per unique description
  ]
}}

Keep each hint to 1 short English sentence. Include ALL {len(unique_descs)} descriptions."""

    learned_mappings = call_claude(client, mappings_prompt, max_tokens=8192)

    return {
        "extraction_hints": extraction_hints,
        "learned_mappings": learned_mappings,
    }


# ── Step 5: Main ────────────────────────────────────────────────────────────

def main():
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        print("ERROR: ANTHROPIC_API_KEY environment variable not set.")
        sys.exit(1)

    fetch_prices = os.environ.get("FETCH_PRICES", "0") == "1"

    print(f"Scanning examples in:\n  {EXAMPLES_BASE}\n")

    # Discover quote files
    cases = find_quote_files()
    print(f"Found {len(cases)} cases:")
    for c in cases:
        print(f"  {c['case'][:50]}")

    # Extract all component rows
    print("\nExtracting component rows...")
    all_rows = []
    case_summaries = []
    for c in cases:
        try:
            rows = extract_quote_rows(c["file"])
            all_rows.extend(rows)
            case_summaries.append({
                "case": c["case"],
                "file": os.path.basename(c["file"]),
                "row_count": len(rows),
            })
            print(f"  {c['case'][:40]:40s} — {len(rows)} rows")
        except Exception as e:
            print(f"  ERROR in {c['case']}: {e}")

    all_descriptions = [r["description"] for r in all_rows]
    print(f"\nTotal rows: {len(all_rows)}")
    print(f"Unique descriptions: {len(set(all_descriptions))}")

    # Optionally fetch price list
    price_list = []
    if fetch_prices:
        price_list = normalize_price_list(fetch_price_list())

    # Analyze with Claude
    print("\nAnalyzing with Claude...")
    try:
        result = analyze_with_claude(all_descriptions, price_list, api_key)
    except Exception as e:
        print(f"ERROR during Claude analysis: {e}")
        sys.exit(1)

    extraction_hints = result.get("extraction_hints", {})
    learned_mappings = result.get("learned_mappings", {})

    # Add metadata
    extraction_hints["_meta"] = {
        "generated": date.today().isoformat(),
        "source_cases": len(cases),
        "unique_descriptions": len(set(all_descriptions)),
    }
    learned_mappings["_meta"] = {
        "generated": date.today().isoformat(),
        "source_cases": len(cases),
        "case_summaries": case_summaries,
    }

    # Write output files
    hints_path = os.path.join(OUT_DIR, "extraction_hints.json")
    mappings_path = os.path.join(OUT_DIR, "learned_mappings.json")

    with open(hints_path, "w", encoding="utf-8") as f:
        json.dump(extraction_hints, f, ensure_ascii=False, indent=2)
    print(f"\nWrote: {hints_path}")

    with open(mappings_path, "w", encoding="utf-8") as f:
        json.dump(learned_mappings, f, ensure_ascii=False, indent=2)
    print(f"Wrote: {mappings_path}")

    # Print summary
    vocab = extraction_hints.get("component_vocabulary", {})
    patterns = learned_mappings.get("patterns", [])
    print(f"\n✅ Done!")
    print(f"   Component categories: {len(vocab)}")
    print(f"   Mapping patterns: {len(patterns)}")
    print(f"\nNext step: Run the runtime integration in backend/main.py")


if __name__ == "__main__":
    main()
