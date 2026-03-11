"""
Import script: Excel price list → Google Sheets
Parses מחירון לוחות י.סופר 10.2025.xlsx and imports with full column support.

Columns imported:
  catalog_number, item_name, unit_price, unit, manufacturer, category, cost, notes

Usage:
    cd "Yoav Sofer/backend"
    pip install openpyxl gspread google-auth
    cd ..
    python scripts/import_pricelist.py
"""

import json
import os
import sys
from pathlib import Path

import gspread
from google.oauth2.service_account import Credentials
from openpyxl import load_workbook

EXCEL_PATH = Path.home() / "Downloads" / "מחירון לוחות י.סופר 10.2025.xlsx"
CREDS_PATH = Path(__file__).parent.parent / "backend" / "config" / "google_credentials.json"
SHEET_ID = "1EckbrWL5jpqLf4Nczq7_b_Euvmq7bExNNQwut5BtXYA"
SHEET_NAME = "מחירון"

WRITE_SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]


def parse_excel(path: Path) -> list[dict]:
    """
    Parse Excel price list.
    Column layout (1-indexed):
        A(1): sequential number (מס"ד)
        B(2): supplier (ספק)
        C(3): catalog number (מק"ט)
        D(4): description (תיאור) — category header OR item name
        E(5): unit (יח' מידה)
        F(6): quantity (כמות)
        G(7): cost/purchase price (עלות)
        H(8): sell price (מחיר)
        I(9): notes (הערות)

    Category row: has A (seq), no B (supplier), no H (price), has D
    Item row: has D and H
    """
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    items = []
    current_category = ""

    for row in ws.iter_rows(min_row=3, values_only=True):  # skip title row (1) and header row (2)
        col_a = row[0] if len(row) > 0 else None
        col_b = row[1] if len(row) > 1 else None
        col_c = row[2] if len(row) > 2 else None
        col_d = row[3] if len(row) > 3 else None
        col_e = row[4] if len(row) > 4 else None
        col_g = row[6] if len(row) > 6 else None
        col_h = row[7] if len(row) > 7 else None
        col_i = row[8] if len(row) > 8 else None

        # Skip truly empty rows
        if col_d is None:
            continue

        col_d_str = str(col_d).strip()
        if not col_d_str:
            continue

        # Determine if price is present
        has_price = col_h is not None and str(col_h).strip() not in ("", "None")
        has_supplier = col_b is not None and str(col_b).strip() not in ("", "None")
        has_seq = col_a is not None and str(col_a).strip() not in ("", "None")

        # Category header detection: has seq, no supplier, no price
        if has_seq and not has_supplier and not has_price:
            current_category = col_d_str
            continue

        # Item row
        if has_price:
            try:
                price = float(str(col_h).replace(",", ".").strip())
            except (ValueError, TypeError):
                price = 0.0

            catalog = str(col_c).strip() if col_c is not None else ""
            if catalog in ("None", ""):
                catalog = ""

            unit = str(col_e).strip() if col_e is not None else "יח'"
            if unit in ("None", ""):
                unit = "יח'"

            manufacturer = str(col_b).strip() if col_b is not None else ""
            if manufacturer in ("None", ""):
                manufacturer = ""

            # cost (עלות) — purchase price, may be numeric or formula string
            cost = ""
            if col_g is not None:
                cost_str = str(col_g).strip()
                if cost_str not in ("", "None"):
                    try:
                        cost = str(round(float(cost_str.replace(",", ".")), 4))
                    except (ValueError, TypeError):
                        cost = cost_str  # keep as-is if formula/text

            notes = str(col_i).strip() if col_i is not None else ""
            if notes in ("None", ""):
                notes = ""

            items.append({
                "catalog_number": catalog,
                "item_name": col_d_str,
                "unit_price": price,
                "unit": unit,
                "manufacturer": manufacturer,
                "category": current_category,
                "cost": cost,
                "notes": notes,
            })

    return items


def import_to_sheets(items: list[dict]) -> None:
    if not CREDS_PATH.exists():
        print(f"ERROR: Credentials not found at {CREDS_PATH}", file=sys.stderr)
        sys.exit(1)

    creds = Credentials.from_service_account_file(str(CREDS_PATH), scopes=WRITE_SCOPES)
    gc = gspread.authorize(creds)
    ss = gc.open_by_key(SHEET_ID)

    try:
        ws = ss.worksheet(SHEET_NAME)
    except gspread.WorksheetNotFound:
        ws = ss.get_worksheet(0)

    # Clear existing content
    ws.clear()

    # Write header row
    header = ["catalog_number", "item_name", "unit_price", "unit", "manufacturer", "category", "cost", "notes"]
    ws.append_row(header)

    # Write all item rows
    rows = [
        [
            item["catalog_number"],
            item["item_name"],
            item["unit_price"],
            item["unit"],
            item["manufacturer"],
            item["category"],
            item.get("cost", ""),
            item.get("notes", ""),
        ]
        for item in items
    ]
    if rows:
        ws.append_rows(rows, value_input_option="USER_ENTERED")

    print(f"✅ Imported {len(rows)} items to Google Sheets '{SHEET_NAME}'")


def main():
    if not EXCEL_PATH.exists():
        print(f"ERROR: Excel file not found at {EXCEL_PATH}", file=sys.stderr)
        sys.exit(1)

    print(f"Parsing {EXCEL_PATH.name}...")
    items = parse_excel(EXCEL_PATH)
    print(f"Found {len(items)} items across categories")

    # Show category summary
    categories: dict[str, int] = {}
    for item in items:
        cat = item["category"] or "ללא קטגוריה"
        categories[cat] = categories.get(cat, 0) + 1
    for cat, count in categories.items():
        print(f"  {cat}: {count} items")

    print(f"\nImporting to Google Sheets (sheet: {SHEET_NAME})...")
    import_to_sheets(items)


if __name__ == "__main__":
    main()
